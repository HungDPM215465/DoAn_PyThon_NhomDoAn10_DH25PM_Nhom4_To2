import tkinter as tk
from tkinter import ttk, messagebox
import mysql.connector
from tkcalendar import DateEntry
from openpyxl import Workbook
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from tkinter import filedialog
import os


DB_CONFIG = {
    "host": "127.0.0.1", 
    "user": "root", 
    "password": "", 
    "database": "qlcuahangtivi" 
}
WINDOW_WIDTH = 1000
WINDOW_HEIGHT = 600


def connect_db(use_database=True):
    """
    Kết nối đến MySQL. 
    use_database=False dùng để kết nối kiểm tra và tạo CSDL.
    """
    config = DB_CONFIG.copy()
    if not use_database:
        config.pop("database", None)
    try:
        return mysql.connector.connect(**config)
    except mysql.connector.Error as err:
        if err.errno == 2003:
             messagebox.showerror("Lỗi Kết Nối CSDL", 
                                  "Lỗi 2003: Không thể kết nối đến MySQL. Vui lòng kiểm tra XAMPP (MySQL đang chạy trên cổng 3306).")
        elif err.errno == mysql.connector.errorcode.ER_ACCESS_DENIED_ERROR:
            messagebox.showerror("Lỗi CSDL", "Tên người dùng hoặc mật khẩu MySQL không đúng.")
        elif err.errno == mysql.connector.errorcode.ER_BAD_DB_ERROR and use_database:
            return None 
        else:
            messagebox.showerror("Lỗi Kết Nối CSDL", f"Kiểm tra cấu hình DB.\nLỗi: {err}")
            return None

def setup_database():
    """Kiểm tra và tạo CSDL cùng các bảng cần thiết."""
    conn = connect_db(use_database=False)
    if conn is None: 
        return False
        
    cur = conn.cursor()
    db_name = DB_CONFIG['database']
    
    try:
        cur.execute(f"CREATE DATABASE IF NOT EXISTS {db_name}")
        conn.database = db_name
        tables = {}
        tables['NhanVien'] = ("CREATE TABLE NhanVien ( MaNV VARCHAR(10) PRIMARY KEY, HoLot VARCHAR(100), Ten VARCHAR(50), ChucVu VARCHAR(50), Phai VARCHAR(10), NgaySinh DATE ) ENGINE=InnoDB")
        tables['KhachHang'] = ("CREATE TABLE KhachHang ( MaKH VARCHAR(10) PRIMARY KEY, TenKH VARCHAR(150), SoDT VARCHAR(15), DiaChi VARCHAR(255) ) ENGINE=InnoDB")
        tables['NhaCungCap'] = ("CREATE TABLE NhaCungCap ( MaNCC VARCHAR(10) PRIMARY KEY, TenNCC VARCHAR(150), SoDTNCC VARCHAR(15), DiaChiNCC VARCHAR(255) ) ENGINE=InnoDB")
        tables['SanPham'] = ("CREATE TABLE SanPham ( MaSP VARCHAR(10) PRIMARY KEY, TenSP VARCHAR(255), DonViTinh VARCHAR(50), GiaBan DECIMAL(10, 2), SoLuongTon INT, MaNCC VARCHAR(10), FOREIGN KEY (MaNCC) REFERENCES NhaCungCap(MaNCC) ON DELETE SET NULL ) ENGINE=InnoDB")
        tables['HoaDon'] = ("CREATE TABLE HoaDon ( MaHD VARCHAR(10) PRIMARY KEY, NgayLap DATE, MaNV VARCHAR(10), MaKH VARCHAR(10), TongTien DECIMAL(10, 2), FOREIGN KEY (MaNV) REFERENCES NhanVien(MaNV) ON DELETE SET NULL, FOREIGN KEY (MaKH) REFERENCES KhachHang(MaKH) ON DELETE SET NULL ) ENGINE=InnoDB")
        tables['ChiTietHoaDon'] = ("CREATE TABLE ChiTietHoaDon ( MaHD VARCHAR(10), MaSP VARCHAR(10), SoLuong INT, DonGia DECIMAL(10, 2), ThanhTien DECIMAL(10, 2), PRIMARY KEY (MaHD, MaSP), FOREIGN KEY (MaHD) REFERENCES HoaDon(MaHD) ON DELETE CASCADE, FOREIGN KEY (MaSP) REFERENCES SanPham(MaSP) ON DELETE CASCADE ) ENGINE=InnoDB")
        tables['LichSuIn'] = ("CREATE TABLE LichSuIn ( id INT AUTO_INCREMENT PRIMARY KEY, MaHD VARCHAR(10), ThoiGianIn DATETIME, NguoiIn VARCHAR(50), FOREIGN KEY (MaHD) REFERENCES HoaDon(MaHD) ON DELETE CASCADE ) ENGINE=InnoDB")

        for name, ddl in tables.items():
            cur.execute(f"SHOW TABLES LIKE '{name}'")
            if not cur.fetchone():
                cur.execute(ddl)
                print(f"Đã tạo bảng: {name}")

        return True
    
    except mysql.connector.Error as err:
        messagebox.showerror("Lỗi CSDL", f"Lỗi tạo bảng: {err}")
        return False
    finally:
        cur.close()
        conn.close()

def create_sample_data():
    """Tự động thêm dữ liệu mẫu nếu bảng còn trống."""
    conn = connect_db()
    if not conn: return

    cur = conn.cursor()
    try:
        # 1. NHÀ CUNG CẤP
        sql_ncc = """
        INSERT IGNORE INTO NhaCungCap (MaNCC, TenNCC, DiaChiNCC, SoDTNCC) VALUES
        ('NCC01', 'Công ty Samsung Vina', 'Số 2, Hải Triều, Q.1, TP.HCM', '02839157310'),
        ('NCC02', 'Sony Electronics VN', 'Tầng 6, Hoàn Kiếm, Hà Nội', '1800588885'),
        ('NCC03', 'LG Electronics VN', 'KCN Nhơn Trạch, Đồng Nai', '18001503');
        """
        cur.execute(sql_ncc)

        # 2. NHÂN VIÊN
        sql_nv = """
        INSERT IGNORE INTO NhanVien (MaNV, HoLot, Ten, Phai, NgaySinh, ChucVu) VALUES
        ('NV01', 'Nguyễn Văn', 'An', 'Nam', '1990-05-15', 'Quản lý'),
        ('NV02', 'Trần Thị', 'Bình', 'Nữ', '1995-08-20', 'Thu ngân'),
        ('NV03', 'Lê Hoàng', 'Nam', 'Nam', '1998-12-10', 'Nhân viên kho');
        """
        cur.execute(sql_nv)

        # 3. KHÁCH HÀNG
        sql_kh = """
        INSERT IGNORE INTO KhachHang (MaKH, TenKH, DiaChi, SoDT) VALUES
        ('KH01', 'Phạm Minh Tuấn', '123 Lê Lợi, Q.1, TP.HCM', '0909123456'),
        ('KH02', 'Đỗ Thúy Hằng', '45 Trần Hưng Đạo, Đà Nẵng', '0918887777'),
        ('KH03', 'Ngô Văn Long', '78 Nguyễn Trãi, Hà Nội', '0987654321');
        """
        cur.execute(sql_kh)

        # 4. SẢN PHẨM (TIVI)
        sql_sp = """
        INSERT IGNORE INTO SanPham (MaSP, TenSP, DonViTinh, GiaBan, SoLuongTon, MaNCC) VALUES
        ('SP001', 'Smart TV Samsung 4K 50 inch', 'Cái', 12500000, 20, 'NCC01'),
        ('SP002', 'Android TV Sony 43 inch', 'Cái', 9800000, 15, 'NCC02'),
        ('SP003', 'TV LG Nanocell 55 inch', 'Cái', 15200000, 10, 'NCC03'),
        ('SP004', 'TV Samsung QLED 65 inch', 'Cái', 22000000, 5, 'NCC01'),
        ('SP005', 'TV Sony OLED 55 inch', 'Cái', 28500000, 8, 'NCC02');
        """
        cur.execute(sql_sp)

        # 5. HÓA ĐƠN
        sql_hd = """
        INSERT IGNORE INTO HoaDon (MaHD, NgayLap, MaNV, MaKH, TongTien) VALUES
        ('HD001', '2023-11-01', 'NV02', 'KH01', 12500000),
        ('HD002', '2023-11-02', 'NV02', 'KH02', 30400000);
        """
        cur.execute(sql_hd)

        # 6. CHI TIẾT HÓA ĐƠN
        sql_cthd = """
        INSERT IGNORE INTO ChiTietHoaDon (MaHD, MaSP, SoLuong, DonGia, ThanhTien) VALUES
        ('HD001', 'SP001', 1, 12500000, 12500000),
        ('HD002', 'SP003', 2, 15200000, 30400000);
        """
        cur.execute(sql_cthd)

        conn.commit()
        print("Đã nạp dữ liệu mẫu thành công!")
    except Exception as e:
        print(f"Lỗi nạp dữ liệu mẫu: {e}")
    finally:
        conn.close()

def center_window(win, w=WINDOW_WIDTH, h=WINDOW_HEIGHT):
    """Canh giữa cửa sổ ứng dụng."""
    ws = win.winfo_screenwidth()
    hs = win.winfo_screenheight()
    x = (ws // 2) - (w // 2)
    y = (hs // 2) - (h // 2)
    win.geometry(f'{w}x{h}+{x}+{y}')



# HÀM MỞ CỬA SỔ CHUNG

def open_window(title, columns, db_table, primary_key, fields_layout):
    """
    Hàm tạo cửa sổ quản lý đa năng.
    - columns: Tên cột trong Database (VD: ['MaNV', 'HoLot'...])
    - fields_layout: Cấu hình giao diện (VD: [("Mã số", 'Entry')...])
    """

    top = tk.Toplevel(root)
    top.title(f"Quản lý {title}")
    center_window(top, 1000, 600)
    top.grab_set() 

    tk.Label(top, text=f"QUẢN LÝ {title.upper()}", font=("Arial", 18, "bold"), fg="#00008B").pack(pady=10)
    
    frame_info = tk.LabelFrame(top, text="Thông tin chi tiết", padx=10, pady=10)
    frame_info.pack(pady=5, padx=10, fill="x")
    
    entries = {}
    widgets = {}

    for i, (label_text, widget_type, *args) in enumerate(fields_layout):
        row, col = i // 2, i % 2 

        tk.Label(frame_info, text=label_text, font=("Arial", 10)).grid(row=row, column=col*2, padx=5, pady=5, sticky="w")

        key = label_text.split(' ')[0]

        if widget_type == 'Entry':
            entry = tk.Entry(frame_info, width=30)
            entry.grid(row=row, column=col*2 + 1, padx=5, pady=5, sticky="w")
            entries[key] = entry
            
        elif widget_type == 'DateEntry':
            date_e = DateEntry(frame_info, width=28, date_pattern="yyyy-mm-dd", background="darkblue", foreground="white") 
            date_e.grid(row=row, column=col*2 + 1, padx=5, pady=5, sticky="w")
            widgets[key] = date_e
            
        elif widget_type == 'Combobox':
            cbb = ttk.Combobox(frame_info, values=args[0], width=28, state="readonly")
            cbb.grid(row=row, column=col*2 + 1, padx=5, pady=5, sticky="w")
            widgets[key] = cbb
            
        elif widget_type == 'Radiobutton':
            var = tk.StringVar(value=args[0][0])
            frame_rb = tk.Frame(frame_info)
            frame_rb.grid(row=row, column=col*2 + 1, padx=5, pady=5, sticky="w")
            for val in args[0]:
                tk.Radiobutton(frame_rb, text=val, variable=var, value=val).pack(side=tk.LEFT, padx=5)
            widgets[key] = var

    frame_table = tk.Frame(top)
    frame_table.pack(padx=10, pady=5, fill="both", expand=True)

    tree = ttk.Treeview(frame_table, columns=columns, show="headings", height=10)

    for i, col_db in enumerate(columns):
        header_text = fields_layout[i][0] if i < len(fields_layout) else col_db
        tree.heading(col_db, text=header_text)
        tree.column(col_db, anchor="center", width=100) 
        if i == 1 or i == 2: tree.column(col_db, width=150, anchor="w")

    scrollbar = ttk.Scrollbar(frame_table, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)


    # LOGIC XỬ LÝ


    def get_input_values():
        """Hàm phụ trợ: Lấy toàn bộ dữ liệu từ form ra một danh sách"""
        vals = []
        for label_text, w_type, *args in fields_layout:
            key = label_text.split(' ')[0]
            val = ""
            if key in entries: val = entries[key].get()
            elif key in widgets:
                w = widgets[key]
                if isinstance(w, DateEntry): val = w.get_date().strftime('%Y-%m-%d')
                else: val = w.get()
            vals.append(val)
        return vals

    def clear_input():
        """Xóa trắng form"""
        for e in entries.values(): 
            e.delete(0, tk.END)
            e.config(state='normal')
        
        for k, w in widgets.items():
            if isinstance(w, DateEntry): w.set_date(datetime.now())
            elif isinstance(w, ttk.Combobox): w.set('')
            elif isinstance(w, tk.StringVar): 
                # Reset Radio về giá trị đầu tiên
                default_val = [item[2][0] for item in fields_layout if 'Radiobutton' in item[1] and item[0].split(' ')[0] == k]
                if default_val: w.set(default_val[0])

        first_key = fields_layout[0][0].split(' ')[0]
        if first_key in entries: entries[first_key].focus()

    def load_data():
        """Tải dữ liệu lên bảng (Đã sửa lỗi isinstance)"""
        for item in tree.get_children(): tree.delete(item)
        
        conn = connect_db()
        if conn:
            try:
                cur = conn.cursor()
                cols_str = ', '.join(columns)
                sql = f"SELECT {cols_str} FROM {db_table}"
                cur.execute(sql)
                rows = cur.fetchall()
                
                for row in rows:
                    display_row = list(row)
                    for idx, val in enumerate(display_row):
                         if val is not None and hasattr(val, 'strftime'):
                            display_row[idx] = val.strftime('%Y-%m-%d')

                    tree.insert("", tk.END, values=display_row)
            except Exception as e:
                messagebox.showerror("Lỗi Tải Dữ liệu", str(e))
            finally:
                conn.close()

    def add_record():
        vals = get_input_values()

        if not vals[0]: 
            messagebox.showwarning("Thiếu dữ liệu", f"Vui lòng nhập {fields_layout[0][0]}!")
            return
            
        conn = connect_db()
        if conn:
            try:
                cur = conn.cursor()

                placeholders = ', '.join(['%s'] * len(columns))
                sql = f"INSERT INTO {db_table} ({', '.join(columns)}) VALUES ({placeholders})"
                
                cur.execute(sql, tuple(vals))
                conn.commit()
                load_data()
                clear_input()
                messagebox.showinfo("Thành công", "Thêm dữ liệu thành công!")
            except mysql.connector.Error as e:
                if e.errno == 1062: 
                    messagebox.showerror("Lỗi Trùng Mã", f"Mã '{vals[0]}' đã tồn tại. Vui lòng nhập mã khác!")
                else:
                    messagebox.showerror("Lỗi CSDL", str(e))
            finally:
                conn.close()

    def select_record(event):
        sel = tree.selection()
        if sel:
            vals = tree.item(sel)['values']
            clear_input()
            for i, (lbl, *_) in enumerate(fields_layout):
                if i >= len(vals): break
                key = lbl.split(' ')[0]
                val = vals[i]
                
                if key in entries: 
                    entries[key].insert(0, val)

                    if i == 0: entries[key].config(state='readonly')
                    
                elif key in widgets:
                    w = widgets[key]
                    if isinstance(w, DateEntry): 
                        if val: w.set_date(str(val))
                    else: w.set(str(val))

    def update_record():
        vals = get_input_values()

        if not vals[0]:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn dòng cần sửa trong bảng trước!")
            return

        conn = connect_db()
        if conn:
            try:
                cur = conn.cursor()
                set_clause = ', '.join([f"{col}=%s" for col in columns[1:]])
                
                params = vals[1:] + [vals[0]]
                
                sql = f"UPDATE {db_table} SET {set_clause} WHERE {primary_key}=%s"
                
                cur.execute(sql, tuple(params))
                conn.commit()
                load_data()
                clear_input()
                messagebox.showinfo("Thành công", "Cập nhật dữ liệu thành công!")
            except Exception as e:
                messagebox.showerror("Lỗi CSDL", str(e))
            finally:
                conn.close()

    def delete_record():
        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Chưa chọn", "Hãy chọn dòng để xóa!")
            return
            
        pk_val = tree.item(sel)['values'][0]

        if messagebox.askyesno("Xác nhận", f"Bạn có chắc chắn muốn xóa mã '{pk_val}' không?"):
            conn = connect_db()
            if conn:
                try:
                    cur = conn.cursor()
                    sql = f"DELETE FROM {db_table} WHERE {primary_key}=%s"
                    cur.execute(sql, (pk_val,))
                    conn.commit()
                    load_data()
                    clear_input()
                    messagebox.showinfo("Thành công", "Đã xóa thành công!")
                except mysql.connector.Error as e:
                    if e.errno == 1451:
                        messagebox.showerror("Không thể xóa", "Dữ liệu này đang được sử dụng ở bảng khác (Ràng buộc khóa ngoại).")
                    else:
                        messagebox.showerror("Lỗi CSDL", str(e))
                finally:
                    conn.close()

    tree.bind('<<TreeviewSelect>>', select_record)

    frame_btn = tk.Frame(top)
    frame_btn.pack(pady=10)

    btn_size = 10
    tk.Button(frame_btn, text="Thêm", width=btn_size, command=add_record, bg="#90ee90").grid(row=0, column=0, padx=5)
    tk.Button(frame_btn, text="Lưu (Sửa)", width=btn_size, command=update_record, bg="#add8e6").grid(row=0, column=1, padx=5)
    tk.Button(frame_btn, text="Xóa", width=btn_size, command=delete_record, bg="#ffcccb").grid(row=0, column=2, padx=5)
    tk.Button(frame_btn, text="Hủy / Mới", width=btn_size, command=clear_input).grid(row=0, column=3, padx=5)
    tk.Button(frame_btn, text="Thoát", width=btn_size, command=top.destroy).grid(row=0, column=4, padx=5)

    load_data()

# QUẢN LÝ HÓA ĐƠN


def open_hoadon_window():
    """Mở cửa sổ lập Hóa đơn bán hàng giống ảnh mẫu."""
    top_hd = tk.Toplevel(root)
    top_hd.title("Hóa đơn bán hàng")
    center_window(top_hd, 1000, 700)
    top_hd.grab_set()

    tk.Label(top_hd, text="HOÁ ĐƠN BÁN HÀNG", font=("Arial", 20, "bold"), fg="blue").pack(pady=10)

    # 1. KHUNG THÔNG TIN CHUNG
    frame_chung = tk.LabelFrame(top_hd, text="Thông tin chung", padx=10, pady=10)
    frame_chung.pack(fill="x", padx=10, pady=5)

    # Biến lưu dữ liệu
    var_ma_hd = tk.StringVar()
    var_ngay_ban = tk.StringVar()
    var_ma_nv = tk.StringVar()
    var_ten_nv = tk.StringVar()
    var_ma_kh = tk.StringVar()
    var_ten_kh = tk.StringVar()
    var_diachi = tk.StringVar()
    var_dienthoai = tk.StringVar()

    current_time = datetime.now()
    var_ma_hd.set(f"HDB{current_time.strftime('%d%m%Y_%H%M')}")

    tk.Label(frame_chung, text="Mã hóa đơn:").grid(row=0, column=0, sticky="w", pady=5)
    tk.Entry(frame_chung, textvariable=var_ma_hd, state="readonly", width=30).grid(row=0, column=1, pady=5)

    tk.Label(frame_chung, text="Ngày bán:").grid(row=1, column=0, sticky="w", pady=5)
    date_ngay_ban = DateEntry(frame_chung, width=27, date_pattern='dd/mm/yyyy')
    date_ngay_ban.grid(row=1, column=1, pady=5)

    tk.Label(frame_chung, text="Mã nhân viên:").grid(row=2, column=0, sticky="w", pady=5)
    cbb_ma_nv = ttk.Combobox(frame_chung, textvariable=var_ma_nv, width=27, state="readonly")
    cbb_ma_nv.grid(row=2, column=1, pady=5)

    tk.Label(frame_chung, text="Tên nhân viên:").grid(row=3, column=0, sticky="w", pady=5)
    tk.Entry(frame_chung, textvariable=var_ten_nv, state="readonly", width=30).grid(row=3, column=1, pady=5)

    tk.Label(frame_chung, text="    ").grid(row=0, column=2) 
    tk.Label(frame_chung, text="Mã khách hàng:").grid(row=0, column=3, sticky="w", pady=5)
    cbb_ma_kh = ttk.Combobox(frame_chung, textvariable=var_ma_kh, width=27, state="readonly")
    cbb_ma_kh.grid(row=0, column=4, pady=5)

    tk.Label(frame_chung, text="Tên khách hàng:").grid(row=1, column=3, sticky="w", pady=5)
    tk.Entry(frame_chung, textvariable=var_ten_kh, state="readonly", width=30).grid(row=1, column=4, pady=5)

    tk.Label(frame_chung, text="Địa chỉ:").grid(row=2, column=3, sticky="w", pady=5)
    tk.Entry(frame_chung, textvariable=var_diachi, state="readonly", width=30).grid(row=2, column=4, pady=5)

    tk.Label(frame_chung, text="Điện thoại:").grid(row=3, column=3, sticky="w", pady=5)
    tk.Entry(frame_chung, textvariable=var_dienthoai, state="readonly", width=30).grid(row=3, column=4, pady=5)

    # 2. KHUNG THÔNG TIN CÁC MẶT HÀNG
    frame_hang = tk.LabelFrame(top_hd, text="Thông tin các mặt hàng", padx=10, pady=10)
    frame_hang.pack(fill="both", expand=True, padx=10, pady=5)

    var_ma_hang = tk.StringVar()
    var_ten_hang = tk.StringVar()
    var_don_gia = tk.DoubleVar()
    var_so_luong = tk.IntVar(value=1)
    var_giam_gia = tk.DoubleVar(value=0)
    var_thanh_tien = tk.DoubleVar(value=0)

    # Dòng 1: Mã hàng, Tên hàng, Đơn giá
    tk.Label(frame_hang, text="Mã hàng:").grid(row=0, column=0, sticky="w", pady=5)
    cbb_ma_hang = ttk.Combobox(frame_hang, textvariable=var_ma_hang, width=15, state="readonly")
    cbb_ma_hang.grid(row=0, column=1, pady=5)

    tk.Label(frame_hang, text="Tên hàng:").grid(row=0, column=2, sticky="w", padx=10)
    tk.Entry(frame_hang, textvariable=var_ten_hang, state="readonly", width=25).grid(row=0, column=3)

    tk.Label(frame_hang, text="Đơn giá:").grid(row=0, column=4, sticky="w", padx=10)
    tk.Entry(frame_hang, textvariable=var_don_gia, state="readonly", width=15).grid(row=0, column=5)

    # Dòng 2: Số lượng, Giảm giá, Thành tiền
    tk.Label(frame_hang, text="Số lượng:").grid(row=1, column=0, sticky="w", pady=5)
    entry_sl = tk.Entry(frame_hang, textvariable=var_so_luong, width=17)
    entry_sl.grid(row=1, column=1)

    tk.Label(frame_hang, text="Giảm giá %:").grid(row=1, column=2, sticky="w", padx=10)
    entry_gg = tk.Entry(frame_hang, textvariable=var_giam_gia, width=25)
    entry_gg.grid(row=1, column=3)

    tk.Label(frame_hang, text="Thành tiền:").grid(row=1, column=4, sticky="w", padx=10)
    tk.Entry(frame_hang, textvariable=var_thanh_tien, state="readonly", width=15).grid(row=1, column=5)

    columns_ct = ("MaHang", "TenHang", "SoLuong", "DonGia", "GiamGia", "ThanhTien")
    tree_ct = ttk.Treeview(frame_hang, columns=columns_ct, show="headings", height=8)
    
    headers = ["Mã hàng", "Tên hàng", "Số lượng", "Đơn giá", "Giảm giá %", "Thành tiền"]
    for col, head in zip(columns_ct, headers):
        tree_ct.heading(col, text=head)
        tree_ct.column(col, width=100, anchor="center")
    tree_ct.column("TenHang", width=200)
    
    tree_ct.grid(row=2, column=0, columnspan=6, pady=10, sticky="nsew")

    # Tổng tiền & Chữ
    frame_total = tk.Frame(frame_hang)
    frame_total.grid(row=3, column=0, columnspan=6, sticky="ew")

    lbl_guide = tk.Label(frame_total, text="Nháy đúp một dòng để xóa", fg="red", font=("Arial", 9, "italic"))
    lbl_guide.pack(side="left")

    var_tong_cong = tk.StringVar(value="0")
    tk.Entry(frame_total, textvariable=var_tong_cong, state="readonly", width=20, font=("Arial", 10, "bold")).pack(side="right")
    tk.Label(frame_total, text="Tổng tiền: ").pack(side="right")

    lbl_bang_chu = tk.Label(frame_hang, text="Bằng chữ: Không đồng", fg="blue", font=("Arial", 10, "bold"))
    lbl_bang_chu.grid(row=4, column=0, columnspan=6, sticky="w", pady=5)

    # 3. KHUNG CÁC NÚT CHỨC NĂNG
    frame_btn = tk.Frame(top_hd)
    frame_btn.pack(pady=10)

    
    # 1. Tải dữ liệu vào Combobox
    def load_combobox_data():
        conn = connect_db()
        if not conn: return
        cur = conn.cursor()
        
        # Load Nhân viên
        cur.execute("SELECT MaNV, Ten FROM NhanVien")
        nvs = cur.fetchall()
        cbb_ma_nv['values'] = [x[0] for x in nvs]

        global map_nv
        map_nv = {x[0]: x[1] for x in nvs}

        # Load Khách hàng
        cur.execute("SELECT MaKH, TenKH, DiaChi, SoDT FROM KhachHang")
        khs = cur.fetchall()
        cbb_ma_kh['values'] = [x[0] for x in khs]
        global map_kh
        map_kh = {x[0]: x for x in khs}

        # Load Hàng hóa
        cur.execute("SELECT MaSP, TenSP, GiaBan FROM SanPham")
        sps = cur.fetchall()
        cbb_ma_hang['values'] = [x[0] for x in sps]
        global map_sp
        map_sp = {x[0]: x for x in sps}
        
        conn.close()

    # 2. Sự kiện chọn Combobox
    def on_select_nv(event):
        ma = cbb_ma_nv.get()
        if ma in map_nv: var_ten_nv.set(map_nv[ma])

    def on_select_kh(event):
        ma = cbb_ma_kh.get()
        if ma in map_kh:
            data = map_kh[ma]
            var_ten_kh.set(data[1])
            var_diachi.set(data[2])
            var_dienthoai.set(data[3])

    def on_select_hang(event):
        ma = cbb_ma_hang.get()
        if ma in map_sp:
            data = map_sp[ma]
            var_ten_hang.set(data[1])
            var_don_gia.set(data[2])
            calculate_thanh_tien()

    # 3. Tính toán tiền
    def calculate_thanh_tien(*args):
        try:
            sl = var_so_luong.get()
            dg = var_don_gia.get()
            gg = var_giam_gia.get()
            tt = (sl * dg) * (1 - gg/100)
            var_thanh_tien.set(tt)
        except:
            var_thanh_tien.set(0)

    var_so_luong.trace("w", calculate_thanh_tien)
    var_giam_gia.trace("w", calculate_thanh_tien)

    # 4. Thêm hàng vào bảng (Cart)
    def add_to_cart():
        ma = var_ma_hang.get()
        if not ma: return
        row = (ma, var_ten_hang.get(), var_so_luong.get(), 
               var_don_gia.get(), var_giam_gia.get(), var_thanh_tien.get())
        tree_ct.insert("", "end", values=row)
        update_total()

    def update_total():
        total = 0
        for item in tree_ct.get_children():
            total += float(tree_ct.item(item, "values")[5])
        var_tong_cong.set(f"{total:,.0f}")
        lbl_bang_chu.config(text=f"Bằng chữ: (Đang cập nhật...)")

    def delete_item_cart(event):
        selected = tree_ct.selection()
        if selected:
            tree_ct.delete(selected)
            update_total()
    
    tree_ct.bind("<Double-1>", delete_item_cart)

    # 5. Lưu hóa đơn xuống CSDL
    def save_invoice():
        if not tree_ct.get_children():
            messagebox.showwarning("Lỗi", "Chưa có mặt hàng nào!")
            return
            
        conn = connect_db()
        if not conn: return
        cur = conn.cursor()
        
        try:
            ma_hd = var_ma_hd.get()
            ngay_lap = date_ngay_ban.get_date()
            cur.execute("INSERT INTO HoaDon (MaHD, NgayLap, MaNV, MaKH, TongTien) VALUES (%s, %s, %s, %s, %s)",
                        (ma_hd, ngay_lap, cbb_ma_nv.get(), cbb_ma_kh.get(), float(var_tong_cong.get().replace(',',''))))

            for item in tree_ct.get_children():
                val = tree_ct.item(item, "values")

                cur.execute("INSERT INTO ChiTietHoaDon (MaHD, MaSP, SoLuong, DonGia, ThanhTien) VALUES (%s, %s, %s, %s, %s)",
                            (ma_hd, val[0], val[2], val[3], val[5]))
                
            conn.commit()
            messagebox.showinfo("Thành công", "Đã lưu hóa đơn!")
            top_hd.destroy()
        except mysql.connector.Error as err:
            messagebox.showerror("Lỗi CSDL", f"Lỗi: {err}")
        finally:
            conn.close()
    

    cbb_ma_nv.bind("<<ComboboxSelected>>", on_select_nv)
    cbb_ma_kh.bind("<<ComboboxSelected>>", on_select_kh)
    cbb_ma_hang.bind("<<ComboboxSelected>>", on_select_hang)

    load_combobox_data()

    btn_add = tk.Button(frame_btn, text="Thêm mặt hàng", width=15, command=add_to_cart, font=("Arial", 9, "bold"))
    btn_add.grid(row=0, column=0, padx=5)

    btn_save = tk.Button(frame_btn, text="Lưu hóa đơn", width=15, command=save_invoice, font=("Arial", 9, "bold"))
    btn_save.grid(row=0, column=1, padx=5)

    btn_cancel = tk.Button(frame_btn, text="Hủy hóa đơn", width=15, command=top_hd.destroy)
    btn_cancel.grid(row=0, column=2, padx=5)
    

    tk.Button(frame_btn, text="In hóa đơn", width=15, state="disabled").grid(row=0, column=3, padx=5)
    tk.Button(frame_btn, text="Đóng", width=15, command=top_hd.destroy).grid(row=0, column=4, padx=5)


    frame_search = tk.Frame(top_hd)
    frame_search.pack(side="bottom", fill="x", pady=10)
    tk.Label(frame_search, text="Mã hóa đơn:").pack(side="left", padx=10)
    tk.Combobox(frame_search, width=20).pack(side="left")
    tk.Button(frame_search, text="Tìm kiếm").pack(side="left", padx=5)




# CHỨC NĂNG ĐẶC BIỆT: TÌM KIẾM SẢN PHẨM


def open_search_product_window():
    """Mở cửa sổ tìm kiếm sản phẩm."""
    
    top_search = tk.Toplevel(root)
    top_search.title("Tìm kiếm Sản phẩm")
    center_window(top_search, 800, 400)
    top_search.grab_set()
    
    lbl_title = tk.Label(top_search, text="TÌM KIẾM SẢN PHẨM (TIVI)", font=("Arial", 14, "bold"))
    lbl_title.pack(pady=10)

    frame_search = tk.Frame(top_search)
    frame_search.pack(pady=5)
    
    tk.Label(frame_search, text="Từ khóa tìm kiếm:").grid(row=0, column=0, padx=5, sticky="w")
    entry_search = tk.Entry(frame_search, width=40)
    entry_search.grid(row=0, column=1, padx=5)

    columns = ("MaSP", "TenSP", "GiaBan", "SoLuongTon", "MaNCC")
    tree_search = ttk.Treeview(top_search, columns=columns, show="headings", height=10)
    for col in columns: tree_search.heading(col, text=col.capitalize())
    tree_search.pack(padx=10, pady=5, fill="both", expand=True)

    def search_product(event=None):
        """Thực hiện tìm kiếm sản phẩm theo tên hoặc mã."""
        keyword = entry_search.get().strip()
        for i in tree_search.get_children(): tree_search.delete(i)
        
        conn = connect_db()
        if conn:
            cur = conn.cursor()
            sql = """
                SELECT MaSP, TenSP, GiaBan, SoLuongTon, MaNCC 
                FROM SanPham 
                WHERE TenSP LIKE %s OR MaSP LIKE %s
            """
            search_param = f"%{keyword}%"
            cur.execute(sql, (search_param, search_param))
            
            for row in cur.fetchall():
                tree_search.insert("", tk.END, values=row)
            
            conn.close()

    tk.Button(frame_search, text="Tìm kiếm", width=15, command=search_product).grid(row=0, column=2, padx=5)
    entry_search.bind('<Return>', search_product)

# BÁO CÁO DOANH THU (THEO THỜI GIAN)

def open_report_revenue_window():
    """Cửa sổ xem báo cáo doanh thu theo khoảng thời gian."""
    top = tk.Toplevel(root)
    top.title("Báo cáo Doanh thu")
    center_window(top, 900, 600)
    top.grab_set()

    tk.Label(top, text="BÁO CÁO DOANH THU", font=("Arial", 18, "bold"), fg="#006400").pack(pady=10)

    # --- Khu vực chọn ngày ---
    frame_filter = tk.Frame(top)
    frame_filter.pack(pady=10)

    tk.Label(frame_filter, text="Từ ngày:").grid(row=0, column=0, padx=5)
    date_from = DateEntry(frame_filter, width=15, date_pattern='yyyy-mm-dd')
    date_from.grid(row=0, column=1, padx=5)
    # Mặc định lấy ngày đầu tháng
    today = datetime.now()
    first_day = today.replace(day=1)
    date_from.set_date(first_day)

    tk.Label(frame_filter, text="Đến ngày:").grid(row=0, column=2, padx=5)
    date_to = DateEntry(frame_filter, width=15, date_pattern='yyyy-mm-dd')
    date_to.grid(row=0, column=3, padx=5)

    # --- Bảng hiển thị ---
    frame_table = tk.Frame(top)
    frame_table.pack(padx=10, fill="both", expand=True)

    cols = ("MaHD", "NgayLap", "TenKH", "NguoiLap", "TongTien")
    tree = ttk.Treeview(frame_table, columns=cols, show="headings", height=15)
    
    headers = ["Mã HĐ", "Ngày Lập", "Khách Hàng", "Nhân Viên", "Tổng Tiền"]
    for col, head in zip(cols, headers):
        tree.heading(col, text=head)
        tree.column(col, anchor="center")
    
    tree.column("TenKH", width=200, anchor="w") # Tên KH rộng hơn
    tree.column("TongTien", anchor="e") # Tiền căn phải

    tree.pack(side="left", fill="both", expand=True)
    scrollbar = ttk.Scrollbar(frame_table, orient="vertical", command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.pack(side="right", fill="y")

    # Label Tổng cộng
    lbl_total = tk.Label(top, text="Tổng doanh thu: 0 VNĐ", font=("Arial", 14, "bold"), fg="red")
    lbl_total.pack(pady=10)

    # --- Hàm lọc dữ liệu ---
    def filter_revenue():
        # Xóa dữ liệu cũ
        for item in tree.get_children(): tree.delete(item)
        
        d_from = date_from.get_date()
        d_to = date_to.get_date()

        conn = connect_db()
        if conn:
            try:
                cur = conn.cursor()
                # Join bảng để lấy tên KH và tên NV thay vì mã
                sql = """
                    SELECT hd.MaHD, hd.NgayLap, kh.TenKH, nv.Ten, hd.TongTien
                    FROM HoaDon hd
                    LEFT JOIN KhachHang kh ON hd.MaKH = kh.MaKH
                    LEFT JOIN NhanVien nv ON hd.MaNV = nv.MaNV
                    WHERE hd.NgayLap BETWEEN %s AND %s
                    ORDER BY hd.NgayLap DESC
                """
                cur.execute(sql, (d_from, d_to))
                rows = cur.fetchall()

                total_revenue = 0
                for row in rows:
                    r = list(row)
                    # Format tiền tệ
                    money_val = r[4] if r[4] else 0
                    total_revenue += money_val
                    r[4] = f"{money_val:,.0f}" # 1,000,000
                    tree.insert("", tk.END, values=r)
                
                # Cập nhật tổng
                lbl_total.config(text=f"Tổng doanh thu: {total_revenue:,.0f} VNĐ")

            except Exception as e:
                messagebox.showerror("Lỗi", str(e))
            finally:
                conn.close()

    tk.Button(frame_filter, text="Xem Báo Cáo", bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), 
              command=filter_revenue).grid(row=0, column=4, padx=10)
    
    # Load mặc định khi mở
    filter_revenue()


# BÁO CÁO TỒN KHO (SẢN PHẨM)

def open_report_inventory_window():
    """Cửa sổ xem tồn kho."""
    top = tk.Toplevel(root)
    top.title("Báo cáo Tồn kho")
    center_window(top, 800, 500)
    top.grab_set()

    tk.Label(top, text="CẢNH BÁO HÀNG TỒN KHO", font=("Arial", 18, "bold"), fg="#FF4500").pack(pady=10)

    frame_table = tk.Frame(top)
    frame_table.pack(padx=10, pady=10, fill="both", expand=True)

    cols = ("MaSP", "TenSP", "DonVi", "GiaBan", "SoLuong")
    tree = ttk.Treeview(frame_table, columns=cols, show="headings")
    
    headers = ["Mã SP", "Tên Sản Phẩm", "Đơn Vị", "Giá Bán", "Số Lượng Tồn"]
    for col, head in zip(cols, headers):
        tree.heading(col, text=head)
        tree.column(col, anchor="center")
    
    tree.column("TenSP", width=250, anchor="w")

    # Tạo tag để tô màu cảnh báo
    tree.tag_configure("low_stock", background="#ffcccc", foreground="red")
    tree.tag_configure("normal", background="white")

    tree.pack(side="left", fill="both", expand=True)
    
    def load_inventory():
        for item in tree.get_children(): tree.delete(item)
        conn = connect_db()
        if conn:
            try:
                cur = conn.cursor()
                # Sắp xếp số lượng tăng dần để thấy hàng sắp hết trước
                cur.execute("SELECT MaSP, TenSP, DonViTinh, GiaBan, SoLuongTon FROM SanPham ORDER BY SoLuongTon ASC")
                rows = cur.fetchall()

                for row in rows:
                    sl = row[4]
                    # Nếu số lượng < 5 thì cảnh báo đỏ
                    tags = ("low_stock",) if sl < 5 else ("normal",)
                    
                    # Format giá tiền
                    display_row = list(row)
                    display_row[3] = f"{row[3]:,.0f}"

                    tree.insert("", tk.END, values=display_row, tags=tags)
            except Exception as e:
                messagebox.showerror("Lỗi", str(e))
            finally:
                conn.close()

    tk.Button(top, text="Tải lại dữ liệu", command=load_inventory).pack(pady=5)
    
    tk.Label(top, text="* Các dòng màu đỏ là hàng sắp hết (Số lượng < 5)", fg="red", font=("Arial", 9, "italic")).pack(pady=5)

    load_inventory()   


# CHỨC NĂNG: IN HÓA ĐƠN RA EXCEL

def export_invoice_to_excel(ma_hd):
    """Xuất chi tiết hóa đơn ra file Excel."""
    conn = connect_db()
    if not conn: return
    
    try:
        cur = conn.cursor()
        
        # 1. Lấy thông tin chung (Hóa đơn, Khách, Nhân viên)
        sql_info = """
            SELECT hd.NgayLap, kh.TenKH, kh.DiaChi, kh.SoDT, nv.Ten, hd.TongTien
            FROM HoaDon hd
            JOIN KhachHang kh ON hd.MaKH = kh.MaKH
            JOIN NhanVien nv ON hd.MaNV = nv.MaNV
            WHERE hd.MaHD = %s
        """
        cur.execute(sql_info, (ma_hd,))
        info = cur.fetchone()
        
        if not info:
            messagebox.showerror("Lỗi", "Không tìm thấy thông tin hóa đơn!")
            return

        # 2. Lấy chi tiết sản phẩm
        sql_details = """
            SELECT sp.TenSP, sp.DonViTinh, ct.SoLuong, ct.DonGia, ct.ThanhTien
            FROM ChiTietHoaDon ct
            JOIN SanPham sp ON ct.MaSP = sp.MaSP
            WHERE ct.MaHD = %s
        """
        cur.execute(sql_details, (ma_hd,))
        details = cur.fetchall()

        # TẠO FILE EXCEL
        wb = Workbook()
        ws = wb.active
        ws.title = "HoaDon"

        # Định dạng Font
        font_bold = Font(bold=True, size=12)
        font_title = Font(bold=True, size=18, color="0000FF")
        border_style = Side(border_style="thin", color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

        # PHẦN HEADER
        ws.merge_cells('A1:E1')
        ws['A1'] = "CỬA HÀNG ĐIỆN TỬ TIVI - LÊ QUỐC HÙNG"
        ws['A1'].font = font_bold
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:E2')
        ws['A2'] = "ĐỊA CHỈ: 123 ĐƯỜNG ABC, TP.HCM - SĐT: 0999.888.777"
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A4:E4')
        ws['A4'] = "HÓA ĐƠN BÁN HÀNG"
        ws['A4'].font = font_title
        ws['A4'].alignment = Alignment(horizontal='center')

        # PHẦN THÔNG TIN KHÁCH HÀNG
        ws['A6'] = f"Mã hóa đơn: {ma_hd}"
        ws['D6'] = f"Ngày lập: {info[0]}"
        
        ws['A7'] = f"Khách hàng: {info[1]}"
        ws['A8'] = f"Địa chỉ: {info[2]}"
        ws['A9'] = f"Điện thoại: {info[3]}"
        ws['D9'] = f"Nhân viên bán: {info[4]}"

        #PHẦN BẢNG SẢN PHẨM
        headers = ["STT", "Tên Sản Phẩm", "ĐVT", "Số Lượng", "Đơn Giá", "Thành Tiền"]
        row_start = 11
        
        # Tạo Header bảng
        ws.cell(row=row_start, column=1, value="STT").font = font_bold
        ws.cell(row=row_start, column=2, value="Tên Sản Phẩm").font = font_bold
        ws.cell(row=row_start, column=3, value="ĐVT").font = font_bold
        ws.cell(row=row_start, column=4, value="Số Lượng").font = font_bold
        ws.cell(row=row_start, column=5, value="Đơn Giá").font = font_bold
        ws.cell(row=row_start, column=6, value="Thành Tiền").font = font_bold
        
        # Kẻ khung Header
        for col in range(1, 7):
            ws.cell(row=row_start, column=col).border = border

        # Ghi dữ liệu sản phẩm
        current_row = row_start + 1
        for i, item in enumerate(details):
            ws.cell(row=current_row, column=1, value=i+1).border = border
            ws.cell(row=current_row, column=2, value=item[0]).border = border
            ws.cell(row=current_row, column=3, value=item[1]).border = border
            ws.cell(row=current_row, column=4, value=item[2]).border = border
            ws.cell(row=current_row, column=5, value=f"{item[3]:,.0f}").border = border
            ws.cell(row=current_row, column=6, value=f"{item[4]:,.0f}").border = border
            current_row += 1

        #PHẦN TỔNG TIỀN
        ws.cell(row=current_row, column=5, value="TỔNG CỘNG:").font = font_bold
        ws.cell(row=current_row, column=6, value=f"{info[5]:,.0f} VNĐ").font = font_bold
        
        #Căn chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40 
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20

        # Lưu file
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")],
                                                 initialfile=f"HoaDon_{ma_hd}.xlsx")
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Thành công", f"Đã xuất hóa đơn ra file:\n{file_path}")
            os.startfile(file_path)

    except Exception as e:
        messagebox.showerror("Lỗi Xuất Excel", str(e))
    finally:
        conn.close()

# GIAO DIỆN LỊCH SỬ HÓA ĐƠN

def open_invoice_history_window():
    """Cửa sổ xem danh sách hóa đơn và in lại."""
    top = tk.Toplevel(root)
    top.title("Lịch sử Hóa đơn")
    center_window(top, 900, 600)
    top.grab_set()

    tk.Label(top, text="LỊCH SỬ HÓA ĐƠN", font=("Arial", 18, "bold"), fg="#8B0000").pack(pady=10)

    # --- Frame bảng ---
    frame_table = tk.Frame(top)
    frame_table.pack(padx=10, pady=5, fill="both", expand=True)

    cols = ("MaHD", "NgayLap", "TenKH", "TenNV", "TongTien")
    tree = ttk.Treeview(frame_table, columns=cols, show="headings", height=15)
    
    headers = ["Mã HĐ", "Ngày Lập", "Khách Hàng", "Nhân Viên Lập", "Tổng Tiền"]
    for col, head in zip(cols, headers):
        tree.heading(col, text=head)
        tree.column(col, anchor="center")
    
    tree.column("TenKH", width=200, anchor="w")
    tree.column("TongTien", anchor="e")

    scrollbar = ttk.Scrollbar(frame_table, orient="vertical", command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    tree.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # --- Load dữ liệu ---
    def load_invoices():
        for item in tree.get_children(): tree.delete(item)
        conn = connect_db()
        if conn:
            try:
                cur = conn.cursor()
                sql = """
                    SELECT hd.MaHD, hd.NgayLap, kh.TenKH, nv.Ten, hd.TongTien
                    FROM HoaDon hd
                    LEFT JOIN KhachHang kh ON hd.MaKH = kh.MaKH
                    LEFT JOIN NhanVien nv ON hd.MaNV = nv.MaNV
                    ORDER BY hd.NgayLap DESC
                """
                cur.execute(sql)
                rows = cur.fetchall()
                for row in rows:
                    r = list(row)
                    r[4] = f"{r[4]:,.0f}"
                    tree.insert("", tk.END, values=r)
            except Exception as e:
                messagebox.showerror("Lỗi", str(e))
            finally:
                conn.close()

    def print_selected_invoice():
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn hóa đơn cần in!")
            return

        ma_hd = tree.item(selected)['values'][0]
        export_invoice_to_excel(ma_hd)


    frame_btn = tk.Frame(top)
    frame_btn.pack(pady=10)
    
    tk.Button(frame_btn, text="In ra Excel", font=("Arial", 11, "bold"), bg="#20B2AA", fg="white", 
              width=15, command=print_selected_invoice).grid(row=0, column=0, padx=10)
    
    tk.Button(frame_btn, text="Làm mới", width=12, command=load_invoices).grid(row=0, column=1, padx=10)
    tk.Button(frame_btn, text="Đóng", width=12, command=top.destroy).grid(row=0, column=2, padx=10)

    load_invoices() 


def create_main_window():
    """Khởi tạo cửa sổ chính với Menu giống ảnh mẫu."""
    global root
    root = tk.Tk()
    root.title("Chương trình quản lý cửa hàng Tivi - Tkinter + MySQL")
    

    root.configure(bg="#A5E6A8") 
    

    center_window(root, 800, 500)
    root.resizable(False, False)


    if not setup_database():
        root.destroy()
        return
    create_sample_data()


    menubar = tk.Menu(root)
    root.config(menu=menubar)

    # 1. Menu: Tập tin
    file_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Tập tin", menu=file_menu)
    file_menu.add_command(label="Đăng xuất", command=lambda: messagebox.showinfo("Thông báo", "Đã đăng xuất!"))
    file_menu.add_separator()
    file_menu.add_command(label="Thoát", command=root.quit)

    # 2. Menu: Danh mục (Quan trọng nhất)
    catalog_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Danh mục", menu=catalog_menu)
    

    layout_nv = [("Mã NV", 'Entry'), ("Họ Lót", 'Entry'), ("Tên", 'Entry'), ("Chức vụ", 'Combobox', ["Quản lý", "NV Bán hàng"]), ("Phái", 'Radiobutton', ["Nam", "Nữ"]), ("Ngày sinh", 'DateEntry')]

    db_cols_nv = ["MaNV", "HoLot", "Ten", "ChucVu", "Phai", "NgaySinh"]
    
    catalog_menu.add_command(label="Nhân viên", command=lambda: open_window("Nhân viên", db_cols_nv, "NhanVien", "MaNV", layout_nv))

    layout_kh = [("Mã KH", 'Entry'), ("Tên KH", 'Entry'), ("Số DT", 'Entry'), ("Địa chỉ", 'Entry')]
    db_cols_kh = ["MaKH", "TenKH", "SoDT", "DiaChi"]
    
    catalog_menu.add_command(label="Khách hàng", command=lambda: open_window("Khách hàng", db_cols_kh, "KhachHang", "MaKH", layout_kh))

    layout_ncc = [("Mã NCC", 'Entry'), ("Tên NCC", 'Entry'), ("Số DTNCC", 'Entry'), ("Địa chỉ NCC", 'Entry')]
    db_cols_ncc = ["MaNCC", "TenNCC", "SoDTNCC", "DiaChiNCC"]
    
    catalog_menu.add_command(label="Nhà cung cấp", command=lambda: open_window("Nhà cung cấp", db_cols_ncc, "NhaCungCap", "MaNCC", layout_ncc))
    
    catalog_menu.add_separator()

    layout_sp = [("Mã SP", 'Entry'), ("Tên SP", 'Entry'), ("Đơn vị tính", 'Entry'), ("Giá bán", 'Entry'), ("Số lượng tồn", 'Entry'), ("Mã NCC", 'Entry')]
    db_cols_sp = ["MaSP", "TenSP", "DonViTinh", "GiaBan", "SoLuongTon", "MaNCC"]

    catalog_menu.add_command(label="Hàng hoá", command=lambda: open_window("Sản phẩm", db_cols_sp, "SanPham", "MaSP", layout_sp))

    bill_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Hoá đơn", menu=bill_menu)
    bill_menu.add_command(label="Lập hoá đơn bán hàng", command=open_hoadon_window)
    bill_menu.add_separator()
    bill_menu.add_command(label="Lịch sử hóa đơn (In lại)", command=open_invoice_history_window)

    # 4. Menu: Tìm kiếm
    search_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Tìm kiếm", menu=search_menu)
    search_menu.add_command(label="Tìm kiếm hàng hoá", command=open_search_product_window)

    # 5. Menu: Báo cáo (ĐÃ CẬP NHẬT)
    report_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Báo cáo", menu=report_menu)
    
    report_menu.add_command(label="Doanh thu bán hàng", command=open_report_revenue_window)
    report_menu.add_separator()
    report_menu.add_command(label="Cảnh báo Tồn kho", command=open_report_inventory_window)

    # 6. Menu: Trợ giúp
    help_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Trợ giúp", menu=help_menu)
    help_menu.add_command(label="Thông tin", command=lambda: messagebox.showinfo("Giới thiệu", "Phần mềm quản lý cửa hàng Tivi\nSinh viên: Lê Quốc Hùng"))

    content_frame = tk.Frame(root, bg="#A5E6A8")
    content_frame.pack(expand=True)

    lbl1 = tk.Label(content_frame, text="Chương trình quản lý", 
                    font=("Arial", 30), fg="#3b65bf", bg="#A5E6A8")
    lbl1.pack(pady=10)

    lbl2 = tk.Label(content_frame, text="CỬA HÀNG TIVI", 
                    font=("Arial", 45, "bold"), fg="#8b2323", bg="#A5E6A8")
    lbl2.pack(pady=10)

    # Dòng tên sinh viên
    lbl3 = tk.Label(root, text="DPM215465 - LÊ QUỐC HÙNG", 
                    font=("Arial", 10, "italic"), fg="#333", bg="#A5E6A8")
    lbl3.pack(side="bottom", pady=20)

    root.mainloop()

if __name__ == "__main__":
    create_main_window()