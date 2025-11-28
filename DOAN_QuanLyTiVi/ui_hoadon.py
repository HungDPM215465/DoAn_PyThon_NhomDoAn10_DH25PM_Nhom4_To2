import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import database as db
import global_state
from datetime import datetime
import os


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

class QuanLyHoaDon(ttk.Frame):
    def __init__(self, parent, role='admin', username='admin', nav_callbacks=None):
        super().__init__(parent)
        self.pack(fill=tk.BOTH, expand=True)
        self.role = role
        self.username = username
        self.nav_callbacks = nav_callbacks 

        COLOR_HEADER = "#2E7D32" 
        COLOR_BG = "#E8F5E9"    
        FONT_LABEL = ("Segoe UI", 11)
        
        style = ttk.Style()
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=30)
        style.configure("Treeview.Heading", font=("Segoe UI", 11, "bold"), foreground="#333")
        style.configure("TNotebook", background=COLOR_BG)

        # 1. HEADER
        header_frame = tk.Frame(self, bg=COLOR_HEADER, height=70)
        header_frame.pack(fill=tk.X)
        
        title_text = "QUẢN LÝ DOANH THU" if role == 'admin' else "GIỎ HÀNG & THANH TOÁN"
        icon = "💲"
        tk.Label(header_frame, text=f"{icon} {title_text}", 
                 font=("Segoe UI", 22, "bold"), bg=COLOR_HEADER, fg="white").pack(side=tk.LEFT, padx=20, pady=15)

        # 2. THANH ĐIỀU HƯỚNG (CHỈ ADMIN)
        if self.role == 'admin' and self.nav_callbacks:
            self.tao_thanh_dieu_huong()

        # 3. BODY (TABs)
        body_frame = tk.Frame(self, bg=COLOR_BG)
        body_frame.pack(fill=tk.BOTH, expand=True)

        self.notebook = ttk.Notebook(body_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Tab 1: Lập Hóa Đơn
        self.tab_tao = tk.Frame(self.notebook, bg=COLOR_BG)
        self.notebook.add(self.tab_tao, text="Lập Hóa Đơn Mới")
        self.setup_tab_tao(COLOR_BG, FONT_LABEL)

        # Tab 2: Lịch Sử (Chỉ Admin)
        if self.role == 'admin':
            self.tab_lichsu = tk.Frame(self.notebook, bg=COLOR_BG)
            self.notebook.add(self.tab_lichsu, text="Lịch Sử Giao Dịch")
            self.setup_tab_lichsu(COLOR_BG)

    def tao_thanh_dieu_huong(self):
        nav_frame = tk.Frame(self, bg="#E0E0E0", height=50)
        nav_frame.pack(fill=tk.X)
        tk.Label(nav_frame, text="Chuyển nhanh: ", bg="#E0E0E0", font=("Segoe UI", 11)).pack(side=tk.LEFT, padx=10)

        def btn_nav(text, func_key, color):
            cmd = lambda: [self.master.destroy(), self.nav_callbacks[func_key]()]
            tk.Button(nav_frame, text=text, command=cmd, bg=color, fg="white", 
                      font=("Segoe UI", 10, "bold"), bd=0, padx=15, pady=5, cursor="hand2").pack(side=tk.LEFT, padx=3, pady=5)

        btn_nav("SẢN PHẨM", 'tivi', "#1A237E")
        btn_nav("NHÂN VIÊN", 'nv', "#EF6C00")
        btn_nav("KHÁCH HÀNG", 'kh', "#6A1B9A")
        btn_nav("NHÀ CUNG CẤP", 'ncc', "#00695C")
        
        tk.Button(nav_frame, text="🏠 VỀ TRANG CHỦ", 
                  command=lambda: [self.master.destroy(), self.nav_callbacks['menu']()],
                  bg="#333", fg="white", font=("Segoe UI", 10, "bold"), bd=0, padx=15, pady=5).pack(side=tk.RIGHT, padx=10)


    def setup_tab_tao(self, bg_color, font_lbl):
        # Frame Chọn Khách
        frame_top = tk.LabelFrame(self.tab_tao, text="Thông tin khách hàng", font=("Segoe UI", 12, "bold"), bg=bg_color, fg="#2E7D32")
        frame_top.pack(fill=tk.X, padx=10, pady=5)
        
        if self.role == 'admin':
            tk.Label(frame_top, text="Chọn Khách:", bg=bg_color, font=font_lbl).pack(side=tk.LEFT, padx=10)
            self.cbo_khach = ttk.Combobox(frame_top, width=35, state="readonly", font=font_lbl)
            self.cbo_khach.pack(side=tk.LEFT, padx=5, pady=10)
            tk.Button(frame_top, text="Tải lại DS", command=self.load_combobox_data, bg="#00897B", fg="white", bd=0, padx=10).pack(side=tk.LEFT, padx=10)
        else:
            lbl_hello = tk.Label(frame_top, text=f"Xin chào: {self.username}", 
                                 font=("Segoe UI", 12, "bold"), bg=bg_color, fg="#2E7D32")
            lbl_hello.pack(side=tk.LEFT, padx=10, pady=10)
            self.id_kh_hien_tai = self.lay_id_khach_hang_tu_user(self.username)
            if not self.id_kh_hien_tai:
                tk.Label(frame_top, text="(Lần đầu mua sắm - Hệ thống sẽ tự tạo hồ sơ)", fg="gray", bg=bg_color).pack(side=tk.LEFT)

        # Frame Thêm Sản Phẩm
        frame_add = tk.LabelFrame(self.tab_tao, text="Thêm sản phẩm", font=("Segoe UI", 12, "bold"), bg=bg_color, fg="#2E7D32")
        frame_add.pack(fill=tk.X, padx=10, pady=5)

        tk.Label(frame_add, text="Sản phẩm:", bg=bg_color, font=font_lbl).grid(row=0, column=0, padx=10, pady=10)
        self.cbo_tivi = ttk.Combobox(frame_add, width=40, state="readonly", font=font_lbl)
        self.cbo_tivi.grid(row=0, column=1, padx=5)
        self.cbo_tivi.bind("<<ComboboxSelected>>", self.on_tivi_select)

        tk.Label(frame_add, text="Giá:", bg=bg_color, font=font_lbl).grid(row=0, column=2, padx=10)
        self.lbl_gia = tk.Label(frame_add, text="0", fg="#D32F2F", font=("Segoe UI", 12, "bold"), bg=bg_color)
        self.lbl_gia.grid(row=0, column=3, padx=5)

        tk.Label(frame_add, text="Số lượng:", bg=bg_color, font=font_lbl).grid(row=0, column=4, padx=10)
        self.spin_sl = ttk.Spinbox(frame_add, from_=1, to=100, width=5, font=font_lbl)
        self.spin_sl.set(1)
        self.spin_sl.grid(row=0, column=5, padx=5)

        tk.Button(frame_add, text="THÊM VÀO GIỎ", command=self.them_vao_gio, bg="#2E7D32", fg="white", bd=0, font=("Segoe UI", 10, "bold"), padx=15).grid(row=0, column=6, padx=20)

        # Bảng Giỏ hàng
        frame_cart = tk.LabelFrame(self.tab_tao, text="Giỏ hàng hiện tại", font=("Segoe UI", 12, "bold"), bg=bg_color, fg="#2E7D32")
        frame_cart.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        cols = ("ID_TV", "Tên TV", "Đơn giá", "Số lượng", "Thành tiền")
        self.tree_cart = ttk.Treeview(frame_cart, columns=cols, show="headings", height=8)
        for c in cols: self.tree_cart.heading(c, text=c)
        self.tree_cart.column("ID_TV", width=50, anchor="center")
        self.tree_cart.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Bottom: Tổng tiền & Nút thanh toán
        frame_bottom = tk.Frame(self.tab_tao, bg=bg_color)
        frame_bottom.pack(fill=tk.X, padx=10, pady=10)

        self.lbl_tongtien = tk.Label(frame_bottom, text="TỔNG TIỀN: 0 VNĐ", font=("Segoe UI", 16, "bold"), fg="#D32F2F", bg=bg_color)
        self.lbl_tongtien.pack(side=tk.RIGHT, padx=20)

        btn_text = "XÁC NHẬN THANH TOÁN" if self.role == 'customer' else "LƯU HÓA ĐƠN"
        btn_color = "#FF5722" if self.role == 'customer' else "#1976D2"
        
        tk.Button(frame_bottom, text=btn_text, font=("Segoe UI", 12, "bold"), bg=btn_color, fg="white", bd=0, padx=20, pady=10, command=self.thanh_toan).pack(side=tk.RIGHT)
        tk.Button(frame_bottom, text="Xóa dòng chọn", command=self.xoa_khoi_gio, bg="#546E7A", fg="white", bd=0, padx=15, pady=10).pack(side=tk.LEFT)

        # Dữ liệu khởi tạo
        self.gio_hang = []; self.map_kh = {}; self.map_tv = {}
        self.load_combobox_data()
        self.nap_tu_global_state()


    def setup_tab_lichsu(self, bg_color):
        paned = ttk.PanedWindow(self.tab_lichsu, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)
        
        # Cột Trái: Danh sách Hóa Đơn
        f_left = tk.LabelFrame(paned, text="Danh sách Hóa Đơn", bg=bg_color, font=("Segoe UI", 11, "bold"), fg="#2E7D32")
        paned.add(f_left, weight=1)
        
        self.tree_hd = ttk.Treeview(f_left, columns=("ID", "Khách", "Ngày", "Tổng"), show="headings")
        for c in ("ID", "Khách", "Ngày", "Tổng"): self.tree_hd.heading(c, text=c)
        self.tree_hd.column("ID", width=50, anchor="center")
        self.tree_hd.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.tree_hd.bind("<<TreeviewSelect>>", self.on_hoadon_select)
        
        # Nút chức năng lịch sử
        btn_frame = tk.Frame(f_left, bg=bg_color)
        btn_frame.pack(pady=5)
        tk.Button(btn_frame, text="Làm mới", command=self.load_lich_su, bg="#00897B", fg="white", bd=0, padx=10).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🖨 Xuất Excel", command=self.in_lai_hoa_don_lich_su, bg="#FF9800", fg="white", bd=0, padx=10).pack(side=tk.LEFT, padx=5)

        # Cột Phải: Chi tiết Hóa Đơn
        f_right = tk.LabelFrame(paned, text="Chi tiết sản phẩm", bg=bg_color, font=("Segoe UI", 11, "bold"), fg="#2E7D32")
        paned.add(f_right, weight=2)
        
        self.tree_ct = ttk.Treeview(f_right, columns=("Tên", "SL", "Giá", "Thành tiền"), show="headings")
        for c in ("Tên", "SL", "Giá", "Thành tiền"): self.tree_ct.heading(c, text=c)
        self.tree_ct.column("SL", width=50, anchor="center")
        self.tree_ct.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.load_lich_su()


    def lay_id_khach_hang_tu_user(self, user):
        conn = db.create_connection(); cursor = conn.cursor()
        cursor.execute("SELECT id_kh FROM khach_hang WHERE email LIKE ?", (f"%{user}%",))
        row = cursor.fetchone(); conn.close()
        return row[0] if row else None

    def nap_tu_global_state(self):
        if global_state.gio_hang_chung:
            for item in global_state.gio_hang_chung:
                self.gio_hang.append(item)
                self.tree_cart.insert("", "end", values=(item['id_tv'], item['ten_tv'], f"{item['don_gia']:,.0f}", item['so_luong'], f"{item['thanh_tien']:,.0f}"))
            self.cap_nhat_tong_tien()

    def load_combobox_data(self):
        self.map_tv = {}
        tv_list = []
        for tv in db.lay_danh_sach_tivi():
            lbl = f"{tv['ten_tv']} ({tv['model']})"
            self.map_tv[lbl] = {'id': tv['id_tv'], 'gia': tv['gia_ban'], 'ton': tv['so_luong_ton']}
            tv_list.append(lbl)
        self.cbo_tivi['values'] = tv_list

        if self.role == 'admin':
            self.map_kh = {}; kh_list = []
            for kh in db.lay_danh_sach_khach_hang():
                lbl = f"{kh['ten_kh']} - {kh['so_dien_thoai']}"
                self.map_kh[lbl] = kh['id_kh']; kh_list.append(lbl)
            self.cbo_khach['values'] = kh_list

    def on_tivi_select(self, event):
        name = self.cbo_tivi.get()
        if name in self.map_tv:
            self.lbl_gia.config(text=f"{self.map_tv[name]['gia']:,.0f}")

    def them_vao_gio(self):
        name = self.cbo_tivi.get()
        if not name: return
        try: sl = int(self.spin_sl.get())
        except: return
        info = self.map_tv[name]
        if sl > info['ton']: messagebox.showwarning("Kho", "Không đủ hàng!"); return
        total = sl * info['gia']
        self.gio_hang.append({'id_tv': info['id'], 'ten_tv': name, 'don_gia': info['gia'], 'so_luong': sl, 'thanh_tien': total})
        self.tree_cart.insert("", "end", values=(info['id'], name, f"{info['gia']:,.0f}", sl, f"{total:,.0f}"))
        self.cap_nhat_tong_tien()

    def xoa_khoi_gio(self):
        sel = self.tree_cart.selection()
        if sel:
            idx = self.tree_cart.index(sel[0]); del self.gio_hang[idx]; self.tree_cart.delete(sel)
            self.cap_nhat_tong_tien(); global_state.gio_hang_chung = self.gio_hang

    def cap_nhat_tong_tien(self):
        total = sum(item['thanh_tien'] for item in self.gio_hang)
        self.lbl_tongtien.config(text=f"TỔNG TIỀN: {total:,.0f} VNĐ")


    def xuat_hoa_don_excel(self, hoa_don_info, danh_sach_sp):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Hóa Đơn"
            
            # Font & Style
            font_title = Font(name='Arial', size=16, bold=True, color="006400")
            font_header = Font(name='Arial', size=11, bold=True)
            font_text = Font(name='Arial', size=11)
            align_center = Alignment(horizontal='center', vertical='center')
            align_right = Alignment(horizontal='right')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Thông tin Cửa hàng
            ws.merge_cells('A1:E1'); ws['A1'] = "CỬA HÀNG ĐIỆN TỬ TIVI STORE"
            ws['A1'].font = Font(name='Arial', size=20, bold=True, color="2E7D32"); ws['A1'].alignment = align_center

            ws.merge_cells('A2:E2'); ws['A2'] = "Địa chỉ: 123 Đường ABC, TP. Long Xuyên, An Giang"
            ws['A2'].alignment = align_center
            
            ws.merge_cells('A3:E3'); ws['A3'] = "Hotline: 1900 1000 - Website: tivistore.com.vn"
            ws['A3'].font = Font(italic=True); ws['A3'].alignment = align_center

            # Tiêu đề Hóa đơn
            ws.merge_cells('A5:E5'); ws['A5'] = "HÓA ĐƠN THANH TOÁN"
            ws['A5'].font = Font(size=18, bold=True); ws['A5'].alignment = align_center

            ws.merge_cells('A6:E6'); ws['A6'] = f"(Số: #{hoa_don_info.get('id_hd', 'New')} - Ngày: {hoa_don_info['ngay_lap']})"
            ws['A6'].alignment = align_center

            # Thông tin Khách
            ws['A8'] = f"Khách hàng: {hoa_don_info['ten_kh']}"
            ws['A9'] = f"Điện thoại: {hoa_don_info['sdt']}"
            ws['A10'] = f"Địa chỉ: {hoa_don_info.get('dia_chi', 'Tại cửa hàng')}"
            for r in range(8, 11): ws[f'A{r}'].font = font_text

            # Bảng Sản phẩm
            headers = ["STT", "Tên Sản Phẩm", "Số Lượng", "Đơn Giá", "Thành Tiền"]
            ws.append([]); ws.append(headers)
            header_row = 12
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=header_row, column=col); c.font = font_header
                c.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                c.border = thin_border; c.alignment = align_center

            curr = 13
            for idx, item in enumerate(danh_sach_sp, 1):
                ws.cell(row=curr, column=1, value=idx).border = thin_border
                ws.cell(row=curr, column=1).alignment = align_center
                ws.cell(row=curr, column=2, value=item['ten_tv']).border = thin_border
                ws.cell(row=curr, column=3, value=item['so_luong']).border = thin_border
                ws.cell(row=curr, column=3).alignment = align_center
                
                # Giá & Thành tiền
                gia = item.get('don_gia', item.get('don_gia_luc_ban', 0))
                c_gia = ws.cell(row=curr, column=4, value=gia)
                c_gia.number_format = '#,##0'; c_gia.border = thin_border
                
                c_tien = ws.cell(row=curr, column=5, value=item['thanh_tien'])
                c_tien.number_format = '#,##0'; c_tien.border = thin_border
                curr += 1

            # Tổng cộng
            ws.merge_cells(f'A{curr}:D{curr}'); ws[f'A{curr}'] = "TỔNG CỘNG:"
            ws[f'A{curr}'].font = font_header; ws[f'A{curr}'].alignment = align_right
            c_tot = ws.cell(row=curr, column=5, value=hoa_don_info['tong_tien'])
            c_tot.font = Font(size=12, bold=True, color="D32F2F"); c_tot.number_format = '#,##0 "VNĐ"'; c_tot.border = thin_border

            # Chân trang
            f_row = curr + 2
            ws[f'B{f_row}'] = "Người mua hàng"; ws[f'D{f_row}'] = "Người lập phiếu"
            ws[f'B{f_row}'].font = Font(bold=True); ws[f'D{f_row}'].font = Font(bold=True)
            ws[f'B{f_row}'].alignment = align_center; ws[f'D{f_row}'].alignment = align_center
            
            # Resize cột
            ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 10; ws.column_dimensions['D'].width = 15; ws.column_dimensions['E'].width = 18

            # Lưu file
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")],
                title="Lưu hóa đơn", initialfile=f"HoaDon_{hoa_don_info.get('id_hd', 'Moi')}_{datetime.now().strftime('%H%M%S')}.xlsx"
            )
            if filename:
                wb.save(filename)
                os.startfile(filename) # Mở file ngay lập tức
                return True
            return False
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi xuất Excel: {e}")
            return False

    def thanh_toan(self):
        if not self.gio_hang: messagebox.showwarning("Trống", "Giỏ hàng đang trống!"); return
        id_kh = None
        ten_kh_display = ""; sdt_kh_display = ""

        if self.role == 'admin':
            txt = self.cbo_khach.get()
            if not txt: messagebox.showwarning("Thiếu", "Chọn khách hàng!"); return
            id_kh = self.map_kh[txt]
            parts = txt.split(" - ")
            ten_kh_display = parts[0]; sdt_kh_display = parts[1] if len(parts) > 1 else ""
        else:
            if self.id_kh_hien_tai: 
                id_kh = self.id_kh_hien_tai; ten_kh_display = self.username; sdt_kh_display = "Đang cập nhật"
            else: 
                db.them_khach_hang(self.username, "000", "Online", f"{self.username}@mail.com")
                kh_moi = db.lay_danh_sach_khach_hang()[-1]
                id_kh = kh_moi['id_kh']; ten_kh_display = kh_moi['ten_kh']; sdt_kh_display = kh_moi['so_dien_thoai']
        
        if messagebox.askyesno("Xác nhận", "Thanh toán hóa đơn này?"):
            if db.tao_hoa_don_moi(id_kh, self.gio_hang):
                messagebox.showinfo("Thành công", "Giao dịch thành công!")
                
                # Hỏi in hóa đơn
                if messagebox.askyesno("In hóa đơn", "Bạn có muốn xuất hóa đơn ra file Excel không?"):
                    info = {
                        'id_hd': 'Moi', 
                        'ten_kh': ten_kh_display, 
                        'sdt': sdt_kh_display,
                        'ngay_lap': datetime.now().strftime("%d/%m/%Y %H:%M"),
                        'tong_tien': sum(i['thanh_tien'] for i in self.gio_hang)
                    }
                    self.xuat_hoa_don_excel(info, self.gio_hang)

                # Reset
                self.gio_hang = []; global_state.gio_hang_chung = []
                for r in self.tree_cart.get_children(): self.tree_cart.delete(r)
                self.cap_nhat_tong_tien(); self.load_combobox_data()
                if self.role == 'admin': self.load_lich_su()
            else: messagebox.showerror("Lỗi", "Lỗi CSDL")

    def in_lai_hoa_don_lich_su(self):
        sel = self.tree_hd.focus()
        if not sel: messagebox.showwarning("Chọn", "Vui lòng chọn hóa đơn cần in!"); return
        
        vals = self.tree_hd.item(sel, "values")
        id_hd, ten_kh, ngay_lap, tong_tien_str = vals
        tong_tien = float(tong_tien_str.replace(",", "").replace(".", ""))

        ct = db.lay_chi_tiet_hoa_don(id_hd)
        info = {'id_hd': id_hd, 'ten_kh': ten_kh, 'sdt': 'Tra cứu hệ thống', 'ngay_lap': ngay_lap, 'tong_tien': tong_tien}
        self.xuat_hoa_don_excel(info, ct)

    def load_lich_su(self):
        for r in self.tree_hd.get_children(): self.tree_hd.delete(r)
        for hd in db.lay_danh_sach_hoa_don():
            ngay = hd['ngay_lap'].strftime("%d/%m/%Y %H:%M") if hd['ngay_lap'] else ""
            self.tree_hd.insert("", "end", values=(hd['id_hd'], hd['ten_kh'], ngay, f"{hd['tong_tien']:,.0f}"))

    def on_hoadon_select(self, event):
        sel = self.tree_hd.focus()
        if not sel: return
        id_hd = self.tree_hd.item(sel, "values")[0]
        for r in self.tree_ct.get_children(): self.tree_ct.delete(r)
        for ct in db.lay_chi_tiet_hoa_don(id_hd):
            self.tree_ct.insert("", "end", values=(ct['ten_tv'], ct['so_luong'], f"{ct['don_gia_luc_ban']:,.0f}", f"{ct['thanh_tien']:,.0f}"))